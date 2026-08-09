"""Compare teacher's 20-item labels with my v2.1 labels · compute agreement + kappa."""
import csv
import sys
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding="utf-8") if hasattr(sys.stdout, "reconfigure") else None

TEACHER_XLSX = Path(__file__).resolve().parent.parent / "data" / "labeled" / "teacher_v2_20items_20260809.xlsx"
CALIB_CSV = Path(__file__).resolve().parent.parent / "data" / "labeled" / "v2_calibration_20_items.csv"

# Column mapping · CSV col name → Excel col label (Chinese)
DIM_MAP = {
    'context_familiarity': ('情境特征', {'0':'0','1':'1','2':'2'}),
    'info_complexity':     ('信息呈现', {'0':'0','1':'1','2':'2'}),
    'topic':               ('模块属性', {'mech':'力学','em':'电磁学','thermal':'热学光学原子物理','cross':'跨模块'}),
    'concept_count':       ('知识点数', {str(i):str(i) for i in range(1,6)}),
    'knowledge_depth':     ('知识深度', {'recall':'记忆识别','comprehend':'基础理解','apply':'常规应用','analyze':'综合分析'}),
    'openness':            ('设问开放度', {'closed':'封闭','semi':'半开放','open':'完全开放'}),
    'modeling':            ('建模复杂度', {str(i):str(i) for i in range(4)}),
    'reasoning':           ('推理链长度', {str(i):str(i) for i in range(1,6)}),
    'computation':         ('运算难度', {'0':'无运算仅定性','1':'简单数值运算','2':'中等多步运算','3':'复杂推导运算'}),
    'question_type':       ('题型因素', {'mcq':'选择题','fill':'填空题','solve':'解答题'}),
}


def load_teacher():
    wb = openpyxl.load_workbook(TEACHER_XLSX, data_only=True)
    ws = wb['Sheet1']
    header = [c.value for c in ws[1]]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        idx = int(row[0])
        item = {}
        for h, v in zip(header[1:], row[1:]):
            if h and v is not None:
                item[h] = str(v).strip()
        result[idx] = item
    return result


def load_mine():
    result = {}
    with open(CALIB_CSV, encoding='utf-8') as f:
        for r in csv.DictReader(f):
            idx = int(r['blind_idx'])
            result[idx] = r
    return result


def cohen_kappa(rater1, rater2, categories):
    """Simple Cohen's kappa for two raters on categorical."""
    n = len(rater1)
    assert n == len(rater2)
    if n == 0: return None
    # Observed agreement
    p_o = sum(1 for a,b in zip(rater1,rater2) if a==b) / n
    # Expected agreement
    from collections import Counter
    c1, c2 = Counter(rater1), Counter(rater2)
    p_e = sum((c1[k]/n) * (c2[k]/n) for k in categories)
    if p_e == 1: return 1.0
    return (p_o - p_e) / (1 - p_e)


def linear_weighted_kappa(rater1, rater2, categories_ordered):
    """Linear-weighted kappa for ordinal data."""
    n = len(rater1)
    if n == 0: return None
    cat_idx = {c:i for i,c in enumerate(categories_ordered)}
    K = len(categories_ordered)
    # Build confusion matrix
    conf = [[0]*K for _ in range(K)]
    for a,b in zip(rater1,rater2):
        if a not in cat_idx or b not in cat_idx: continue
        conf[cat_idx[a]][cat_idx[b]] += 1
    # Marginals
    row_sums = [sum(row) for row in conf]
    col_sums = [sum(conf[r][c] for r in range(K)) for c in range(K)]
    tot = sum(row_sums)
    if tot == 0: return None
    # Weights (linear)
    w = [[abs(i-j)/(K-1) for j in range(K)] for i in range(K)]
    # Numerator and denominator
    num = sum(w[i][j] * conf[i][j] for i in range(K) for j in range(K))
    den = sum(w[i][j] * (row_sums[i]*col_sums[j]/tot) for i in range(K) for j in range(K))
    if den == 0: return None
    return 1 - num/den


def main():
    teacher = load_teacher()
    mine = load_mine()

    print(f'Teacher labeled: {len(teacher)} items')
    print(f'Mine labeled:    {len(mine)} items')
    print()

    print('=' * 100)
    print(f'{"维度":15} {"我原打":25} {"杨老师":25} {"一致 (%)":10} {"简单 kappa":12} {"加权 kappa":12}')
    print('=' * 100)

    disagreements = {}

    for dim_key, (chinese_name, value_map) in DIM_MAP.items():
        # Get mine (English values) and teacher (Chinese values), align via value_map
        # For each idx, translate my English to Chinese, compare to teacher Chinese
        mine_translated = []
        teacher_vals = []
        for idx in sorted(teacher.keys()):
            if idx not in mine: continue
            my_eng = str(mine[idx].get(dim_key, '')).strip()
            my_ch = value_map.get(my_eng, my_eng)
            teach_ch = teacher[idx].get(chinese_name, '').strip()
            mine_translated.append(my_ch)
            teacher_vals.append(teach_ch)
            if my_ch != teach_ch:
                disagreements.setdefault(dim_key, []).append((idx, my_ch, teach_ch))

        agree = sum(1 for a,b in zip(mine_translated, teacher_vals) if a==b)
        n = len(mine_translated)
        pct = 100.0*agree/n if n>0 else 0

        # Categories (all Chinese values)
        all_cats = set(mine_translated) | set(teacher_vals)
        # For ordinal, use spec order; else fallback
        if dim_key in ('knowledge_depth', 'openness', 'computation'):
            ordered_cats = list(value_map.values())
        elif dim_key in ('context_familiarity', 'info_complexity', 'concept_count', 'modeling', 'reasoning'):
            ordered_cats = sorted(all_cats, key=lambda x: int(x) if x.isdigit() else 999)
        else:
            ordered_cats = sorted(all_cats)

        simple_k = cohen_kappa(mine_translated, teacher_vals, all_cats)
        weighted_k = linear_weighted_kappa(mine_translated, teacher_vals, ordered_cats) if dim_key in ('context_familiarity','info_complexity','concept_count','modeling','reasoning','knowledge_depth','openness','computation') else None

        w_str = f'{weighted_k:.3f}' if weighted_k is not None else 'n/a'
        k_str = f'{simple_k:.3f}' if simple_k is not None else 'n/a'
        my_dist = f'({",".join(sorted(set(mine_translated)))})'
        teach_dist = f'({",".join(sorted(set(teacher_vals)))})'
        print(f'{chinese_name:15} {my_dist:25} {teach_dist:25} {pct:>7.1f}%   {k_str:>10}   {w_str:>10}')

    print()
    print('=' * 100)
    print('分歧详情 (仅显示不一致的题):')
    print('=' * 100)
    for dim, items in disagreements.items():
        chinese_name = DIM_MAP[dim][0]
        print(f'\n{chinese_name} · 共 {len(items)}/20 条分歧:')
        for idx, mine_v, teach_v in items:
            print(f'  第 {idx} 题: 我打 [{mine_v}] · 老师打 [{teach_v}]')


if __name__ == '__main__':
    main()
