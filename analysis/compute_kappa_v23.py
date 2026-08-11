"""v2.3 kappa · concept_count only · using updated teacher gold (Item 1 corrected 1→2 per 杨老师 15:27)."""
import csv, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
TEACHER = ROOT / "data/labeled/teacher_v23_20items_20260811.xlsx"
MINE = ROOT / "data/labeled/v2_3_calibration_20_items.csv"


def load_teacher():
    wb = openpyxl.load_workbook(TEACHER, data_only=True)
    ws = wb['Sheet1']
    header = [c.value for c in ws[1]]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        idx = int(row[0])
        result[idx] = {h: v for h, v in zip(header[1:], row[1:])}
    return result


def kappa(r1, r2):
    n = len(r1)
    p_o = sum(1 for a, b in zip(r1, r2) if a == b) / n
    cats = set(r1) | set(r2)
    c1, c2 = Counter(r1), Counter(r2)
    p_e = sum((c1[k]/n)*(c2[k]/n) for k in cats)
    return (p_o - p_e) / (1 - p_e) if p_e < 1 else 1.0


def wkappa(r1, r2, cats):
    n = len(r1)
    idx = {c: i for i, c in enumerate(cats)}
    K = len(cats)
    conf = [[0]*K for _ in range(K)]
    for a, b in zip(r1, r2):
        if a in idx and b in idx:
            conf[idx[a]][idx[b]] += 1
    rows = [sum(r) for r in conf]
    cols = [sum(conf[r][c] for r in range(K)) for c in range(K)]
    tot = sum(rows)
    w = [[abs(i-j)/(K-1) for j in range(K)] for i in range(K)]
    num = sum(w[i][j]*conf[i][j] for i in range(K) for j in range(K))
    den = sum(w[i][j]*(rows[i]*cols[j]/tot) for i in range(K) for j in range(K))
    return 1 - num/den if den > 0 else None


def main():
    teacher = load_teacher()
    with open(MINE, encoding='utf-8') as f:
        mine = {int(r['blind_idx']): r for r in csv.DictReader(f)}

    mine_vals, teach_vals, diffs = [], [], []
    for idx in sorted(teacher):
        m = str(mine[idx]['concept_count'])
        t = str(teacher[idx]['知识点数'])
        mine_vals.append(m)
        teach_vals.append(t)
        if m != t:
            diffs.append((idx, m, t))

    agree = sum(1 for m, t in zip(mine_vals, teach_vals) if m == t)
    print(f'v2.3 vs teacher (updated per 杨老师 15:27 Item 1) · 20 items')
    print(f'一致: {agree}/20 = {100*agree/20:.0f}%')
    sk = kappa(mine_vals, teach_vals)
    cats = sorted(set(mine_vals) | set(teach_vals), key=int)
    wk = wkappa(mine_vals, teach_vals, cats)
    print(f'Simple kappa: {sk:.3f}')
    print(f'Weighted kappa: {wk:.3f}')
    print()
    print(f'evolution:')
    print(f'  v2.1 (30%/0.05/0.30) → v2.3 first pass (85%/0.798/0.869) → v2.3 updated (?/?/?)')
    print()
    print(f'剩余分歧:')
    for idx, m, t in diffs:
        print(f'  第{idx}题: AI={m} · 老师={t}')


if __name__ == "__main__":
    main()
